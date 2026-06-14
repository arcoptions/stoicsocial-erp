"""Workbook importer for BoldERP.

Run from Django so models/settings are already loaded:

python manage.py shell -c "from importer import run; run(r'C:\\path\\to\\master-data.xlsx')"

Optional:

python manage.py shell -c "from importer import run; run(r'C:\\path\\to\\master-data.xlsx', dry_run=True)"

Expected sheets:
- vendors: name, contact, is_active
- designs: name, product_type, sub_category, has_variants, variants, notes
- design_assets: design_name, colour, artwork_url, mockup_url, blank_fabric
- blank_skus: fabric, colour, size, on_hand, reserved, reorder_min, reorder_target
- printed_skus: design_name, variant, colour, size, on_hand, reserved, buffer_min, buffer_target, buffer_max

`variants` may be a JSON array string or a comma-separated list.
"""

from __future__ import annotations

import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import openpyxl
from django.db import transaction

from core.models import BlankSKU, Design, DesignAsset, PrintedSKU, Vendor


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _nullable_text(value: Any) -> str | None:
    text = _normalize_text(value)
    return text or None


def _to_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    return int(float(text)) if text else default


def _to_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _to_list(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("["):
        parsed = json.loads(text)
        if not isinstance(parsed, list):
            raise ValueError("Expected JSON array for variants")
        return [str(item).strip() for item in parsed if str(item).strip()]
    return [item.strip() for item in text.split(",") if item.strip()]


def _sheet_rows(workbook: openpyxl.Workbook, sheet_name: str) -> Iterable[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [_normalize_header(value) for value in next(rows)]
    except StopIteration:
        return []
    return [
        {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        for row in rows
        if any(value not in (None, "") for value in row)
    ]


def _import_vendors(workbook: openpyxl.Workbook, summary: dict[str, int]) -> None:
    for row in _sheet_rows(workbook, "vendors"):
        name = _normalize_text(row.get("name"))
        if not name:
            continue
        Vendor.objects.update_or_create(
            name=name,
            defaults={
                "contact": _normalize_text(row.get("contact")),
                "is_active": _to_bool(row.get("is_active"), default=True),
            },
        )
        summary["vendors"] += 1


def _import_designs(workbook: openpyxl.Workbook, summary: dict[str, int]) -> None:
    for row in _sheet_rows(workbook, "designs"):
        name = _normalize_text(row.get("name"))
        if not name:
            continue
        variants = _to_list(row.get("variants"))
        Design.objects.update_or_create(
            name=name,
            defaults={
                "product_type": _normalize_text(row.get("product_type")) or Design.ProductType.TSHIRT,
                "sub_category": _normalize_text(row.get("sub_category")) or Design.SubCategory.REGULAR,
                "has_variants": _to_bool(row.get("has_variants"), default=bool(variants)),
                "variants": variants,
                "notes": _normalize_text(row.get("notes")),
            },
        )
        summary["designs"] += 1


def _import_design_assets(workbook: openpyxl.Workbook, summary: dict[str, int]) -> None:
    for row in _sheet_rows(workbook, "design_assets"):
        design_name = _normalize_text(row.get("design_name"))
        colour = _normalize_text(row.get("colour"))
        if not design_name or not colour:
            continue
        design = Design.objects.get(name=design_name)
        DesignAsset.objects.update_or_create(
            design=design,
            colour=colour,
            defaults={
                "artwork_url": _normalize_text(row.get("artwork_url")),
                "mockup_url": _normalize_text(row.get("mockup_url")),
                "blank_fabric": _normalize_text(row.get("blank_fabric")),
            },
        )
        summary["design_assets"] += 1


def _import_blank_skus(workbook: openpyxl.Workbook, summary: dict[str, int]) -> None:
    for row in _sheet_rows(workbook, "blank_skus"):
        fabric = _normalize_text(row.get("fabric")) or _normalize_text(row.get("fabric_gsm"))
        colour = _normalize_text(row.get("colour"))
        size = _normalize_text(row.get("size"))
        if not fabric or not colour or not size:
            continue
        BlankSKU.objects.update_or_create(
            fabric=fabric,
            colour=colour,
            size=size,
            defaults={
                "on_hand": _to_int(row.get("on_hand", row.get("on_hand_qty"))),
                "reserved": _to_int(row.get("reserved", row.get("reserved_qty"))),
                "reorder_min": _to_int(row.get("reorder_min")),
                "reorder_target": _to_int(row.get("reorder_target")),
            },
        )
        summary["blank_skus"] += 1


def _import_printed_skus(workbook: openpyxl.Workbook, summary: dict[str, int]) -> None:
    for row in _sheet_rows(workbook, "printed_skus"):
        design_name = _normalize_text(row.get("design_name"))
        colour = _normalize_text(row.get("colour"))
        if not design_name or not colour:
            continue
        design = Design.objects.get(name=design_name)
        PrintedSKU.objects.update_or_create(
            design=design,
            variant=_nullable_text(row.get("variant")),
            colour=colour,
            size=_nullable_text(row.get("size")),
            defaults={
                "on_hand": _to_int(row.get("on_hand", row.get("on_hand_qty"))),
                "reserved": _to_int(row.get("reserved", row.get("reserved_qty"))),
                "buffer_min": _to_int(row.get("buffer_min", row.get("min_buffer_qty"))),
                "buffer_target": _to_int(row.get("buffer_target", row.get("target_buffer_qty"))),
                "buffer_max": _to_int(row.get("buffer_max", row.get("max_buffer_qty"))),
            },
        )
        summary["printed_skus"] += 1


@transaction.atomic
def run(workbook_path: str | Path, dry_run: bool = False) -> dict[str, int]:
    """Import workbook sheets into BoldERP models inside a single transaction."""
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = openpyxl.load_workbook(path, data_only=True)
    summary: dict[str, int] = {
        "vendors": 0,
        "designs": 0,
        "design_assets": 0,
        "blank_skus": 0,
        "printed_skus": 0,
    }

    _import_vendors(workbook, summary)
    _import_designs(workbook, summary)
    _import_design_assets(workbook, summary)
    _import_blank_skus(workbook, summary)
    _import_printed_skus(workbook, summary)

    if dry_run:
        transaction.set_rollback(True)

    return summary


def print_summary(summary: dict[str, int]) -> None:
    """Render a compact import summary for shell usage."""
    total = sum(summary.values())
    print(f"Imported rows: {total}")
    for key, value in summary.items():
        print(f"- {key}: {value}")

