from __future__ import annotations

from io import BytesIO
from typing import Any

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db import IntegrityError, transaction
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from core.models import BlankSKU, DeletedInventoryItem, Design, DesignAsset, PrintedSKU, Vendor
from core.security import inventory_access_required, is_admin_user

SIZE_COLUMNS = ["S", "M", "L", "XL", "2XL", "3XL"]


def _is_ajax(request: HttpRequest) -> bool:
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _to_int(value: Any, *, default: int = 0) -> int:
    if value in (None, ""):
        return default
    return int(str(value).strip())


def _normalize_size(value: str | None) -> str:
    normalized = str(value or "").strip().upper()
    if normalized in SIZE_COLUMNS:
        return normalized
    return "NA"


def _build_printed_groups(printed_skus: list[PrintedSKU]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for sku in printed_skus:
        key = (sku.design.name, sku.variant or "", sku.colour)
        row = grouped.setdefault(
            key,
            {
                "design_name": sku.design.name,
                "design_id": sku.design.id,
                "variant": sku.variant or "",
                "colour": sku.colour,
                "design_asset_id": None,
                "design_asset_blank_sku_id": None,
                "design_asset_blank_sku_label": None,
                "linked_plain_option": "",
                "on_hand": {size: 0 for size in SIZE_COLUMNS},
                "reserved": {size: 0 for size in SIZE_COLUMNS},
                "buffer_min": {size: 0 for size in SIZE_COLUMNS},
                "buffer_target": {size: 0 for size in SIZE_COLUMNS},
                "buffer_max": {size: 0 for size in SIZE_COLUMNS},
                "sku_ids": {size: "" for size in SIZE_COLUMNS},
                "total_on_hand": 0,
                "total_reserved": 0,
            },
        )
        # Get design_asset info from the first sku
        if row["design_asset_id"] is None and sku.design_asset:
            row["design_asset_id"] = sku.design_asset.id
            row["design_asset_blank_sku_id"] = str(sku.design_asset.blank_sku_id) if sku.design_asset.blank_sku_id else None
            row["design_asset_blank_sku_label"] = str(sku.design_asset.blank_sku) if sku.design_asset.blank_sku_id else None
            # Prefer explicit blank_sku link; fallback to fabric-level link for consolidated UI.
            if sku.design_asset.blank_sku_id:
                row["linked_plain_option"] = f"{sku.design_asset.blank_sku.fabric}|||{sku.design_asset.blank_sku.colour}"
            elif sku.design_asset.blank_fabric:
                row["linked_plain_option"] = f"{sku.design_asset.blank_fabric}|||{sku.colour}"
        
        size_key = _normalize_size(sku.size)
        if size_key in SIZE_COLUMNS:
            row["on_hand"][size_key] = sku.on_hand
            row["reserved"][size_key] = sku.reserved
            row["buffer_min"][size_key] = sku.buffer_min
            row["buffer_target"][size_key] = sku.buffer_target
            row["buffer_max"][size_key] = sku.buffer_max
            row["sku_ids"][size_key] = str(sku.id)
        row["total_on_hand"] += sku.on_hand
        row["total_reserved"] += sku.reserved

    rows = list(grouped.values())
    for row in rows:
        row["size_cells"] = [
            {
                "size": size,
                "on_hand": row["on_hand"][size],
                "reserved": row["reserved"][size],
                "buffer_min": row["buffer_min"][size],
                "buffer_target": row["buffer_target"][size],
                "buffer_max": row["buffer_max"][size],
                "sku_id": row["sku_ids"][size],
            }
            for size in SIZE_COLUMNS
        ]
    rows.sort(key=lambda r: (r["design_name"].lower(), r["variant"].lower(), r["colour"].lower()))
    return rows


def _build_plain_groups(blank_skus: list[BlankSKU]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for sku in blank_skus:
        key = (sku.fabric, sku.colour)
        row = grouped.setdefault(
            key,
            {
                "fabric": sku.fabric,
                "colour": sku.colour,
                "on_hand": {size: 0 for size in SIZE_COLUMNS},
                "total_on_hand": 0,
                "size_sku_ids": {size: "" for size in SIZE_COLUMNS},
            },
        )
        size_key = _normalize_size(sku.size)
        if size_key in SIZE_COLUMNS:
            row["on_hand"][size_key] = sku.on_hand
            row["size_sku_ids"][size_key] = str(sku.id)
        row["total_on_hand"] += sku.on_hand

    rows = list(grouped.values())
    for row in rows:
        row["size_cells"] = [
            {
                "size": size,
                "on_hand": row["on_hand"][size],
                "sku_id": row["size_sku_ids"][size],
            }
            for size in SIZE_COLUMNS
        ]
    rows.sort(key=lambda r: (r["fabric"].lower(), r["colour"].lower()))
    return rows


@login_required
@inventory_access_required
def sku_manager(request: HttpRequest) -> HttpResponse:
    """Manage plain and printed stock with searchable, editable consolidated views."""
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "create_printed":
            return _create_printed_sku(request)
        if action == "update_printed_group":
            return _update_printed_group(request)
        if action == "delete_printed_group":
            return _delete_printed_group(request)
        if action == "update_plain":
            return _update_plain_sku(request)
        if action == "update_plain_group":
            return _update_plain_sku(request)
        if action == "import_printed":
            return _import_printed_skus(request)
        if action == "link_blank_sku":
            return _link_blank_sku(request)

    q = request.GET.get("q", "").strip()
    colour_filter = request.GET.get("colour", "").strip()
    fabric_filter = request.GET.get("fabric", "").strip()
    stock_type = request.GET.get("stock_type", "printed").strip().lower()
    if stock_type not in {"plain", "printed"}:
        stock_type = "printed"

    blank_qs = BlankSKU.objects.order_by("fabric", "colour", "size")
    if q:
        blank_qs = blank_qs.filter(
            Q(fabric__icontains=q) | Q(colour__icontains=q) | Q(size__icontains=q)
        )
    if fabric_filter:
        blank_qs = blank_qs.filter(fabric__iexact=fabric_filter)

    printed_qs = PrintedSKU.objects.select_related("design", "design_asset", "design_asset__blank_sku").filter(is_active=True).order_by("design__name", "variant", "colour", "size")
    if q:
        printed_qs = printed_qs.filter(
            Q(design__name__icontains=q)
            | Q(variant__icontains=q)
            | Q(colour__icontains=q)
            | Q(size__icontains=q)
        )
    if colour_filter:
        printed_qs = printed_qs.filter(colour__iexact=colour_filter)

    blank_skus = list(blank_qs)
    plain_groups = _build_plain_groups(blank_skus)
    printed_skus = list(printed_qs)
    printed_groups = _build_printed_groups(printed_skus)
    deleted_items = DeletedInventoryItem.objects.filter(restored_at__isnull=True).order_by("-created_at")[:100]
    colours = list(PrintedSKU.objects.order_by("colour").values_list("colour", flat=True).distinct())
    fabrics = list(BlankSKU.objects.order_by("fabric").values_list("fabric", flat=True).distinct())
    # Consolidated linking options: one option per fabric+colour (not per size).
    all_blank_skus = list(
        BlankSKU.objects.order_by("fabric", "colour")
        .values("fabric", "colour")
        .distinct()
    )

    return render(
        request,
        "core/sku_manager.html",
        {
            "blank_skus": blank_skus,
            "plain_groups": plain_groups,
            "printed_groups": printed_groups,
            "size_columns": SIZE_COLUMNS,
            "deleted_items": deleted_items,
            "is_admin": is_admin_user(request.user),
            "colours": colours,
            "fabrics": fabrics,
            "all_blank_skus": all_blank_skus,
            "filters": {
                "q": q,
                "colour": colour_filter,
                "fabric": fabric_filter,
                "stock_type": stock_type,
            },
        },
    )


@login_required
@inventory_access_required
def printed_sku_template(request: HttpRequest) -> HttpResponse:
    """Download Excel template for printed SKU bulk import."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Printed SKUs"
    sheet.append([
        "design_name",
        "variant",
        "colour",
        "size",
        "on_hand",
        "reserved",
        "buffer_min",
        "buffer_target",
        "buffer_max",
    ])
    sheet.append([
        "Urban Eagle",
        "Oversized",
        "Black",
        "M",
        25,
        2,
        5,
        20,
        35,
    ])

    data = BytesIO()
    workbook.save(data)
    data.seek(0)

    response = HttpResponse(
        data.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="printed_sku_import_template.xlsx"'
    return response


@login_required
@inventory_access_required
@transaction.atomic
def restore_deleted_item(request: HttpRequest, item_id: str) -> HttpResponse:
    """Restore a deleted SKU record (admin only)."""
    if not is_admin_user(request.user):
        messages.error(request, "Only admin users can restore deleted items.")
        return redirect("sku-manager")

    if request.method != "POST":
        return redirect("sku-manager")

    item = DeletedInventoryItem.objects.select_for_update().filter(id=item_id).first()
    if item is None or item.restored_at is not None:
        messages.error(request, "Deleted item not found or already restored.")
        return redirect("sku-manager")

    payload = item.payload
    if item.record_type == DeletedInventoryItem.RecordType.PRINTED_SKU:
        design_name = str(payload.get("design_name", "")).strip()
        if not design_name:
            messages.error(request, "Cannot restore printed SKU: missing design name.")
            return redirect("sku-manager")
        design, _ = Design.objects.get_or_create(name=design_name)
        PrintedSKU.objects.update_or_create(
            design=design,
            variant=(str(payload.get("variant", "")).strip() or None),
            colour=str(payload.get("colour", "")).strip(),
            size=(str(payload.get("size", "")).strip() or None),
            defaults={
                "on_hand": _to_int(payload.get("on_hand"), default=0),
                "reserved": _to_int(payload.get("reserved"), default=0),
                "is_active": True,
                "buffer_min": _to_int(payload.get("buffer_min"), default=0),
                "buffer_target": _to_int(payload.get("buffer_target"), default=0),
                "buffer_max": _to_int(payload.get("buffer_max"), default=0),
            },
        )
    else:
        BlankSKU.objects.update_or_create(
            fabric=str(payload.get("fabric", "")).strip(),
            colour=str(payload.get("colour", "")).strip(),
            size=str(payload.get("size", "")).strip(),
            defaults={
                "on_hand": _to_int(payload.get("on_hand"), default=0),
                "reserved": _to_int(payload.get("reserved"), default=0),
                "reorder_min": _to_int(payload.get("reorder_min"), default=0),
                "reorder_target": _to_int(payload.get("reorder_target"), default=0),
            },
        )

    item.restored_at = timezone.now()
    item.save(update_fields=["restored_at", "updated_at"])
    messages.success(request, f"Restored: {item.label}")
    return redirect("sku-manager")


@transaction.atomic
def _update_plain_sku(request: HttpRequest) -> HttpResponse:
    """Update one plain SKU row with locking for consistency."""
    sku_id = request.POST.get("blank_sku_id", "").strip()
    sku = BlankSKU.objects.select_for_update().filter(id=sku_id).first()
    if sku is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": "Plain SKU not found."}, status=404)
        messages.error(request, "Plain SKU not found.")
        return redirect("sku-manager")

    try:
        sku.on_hand = _to_int(request.POST.get("on_hand"), default=0)
        sku.save(update_fields=["on_hand", "updated_at"])
    except ValueError as exc:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": f"Could not update plain SKU: {exc}"}, status=400)
        messages.error(request, f"Could not update plain SKU: {exc}")
        return redirect("sku-manager")

    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": "Saved"}, status=200)

    messages.success(request, "Plain SKU updated.")
    return redirect("sku-manager")


@transaction.atomic
def _create_printed_sku(request: HttpRequest) -> HttpResponse:
    """Create a new printed SKU row from form input."""
    design_name = request.POST.get("design_name", "").strip()
    colour = request.POST.get("colour", "").strip()
    if not design_name or not colour:
        messages.error(request, "Design name and colour are required.")
        return redirect("sku-manager")

    design, _ = Design.objects.get_or_create(name=design_name)
    try:
        variant = request.POST.get("variant", "").strip() or None
        size = request.POST.get("size", "").strip() or None
        existing = PrintedSKU.objects.select_for_update().filter(
            design=design,
            variant=variant,
            colour=colour,
            size=size,
        ).first()
        if existing is None:
            PrintedSKU.objects.create(
                design=design,
                variant=variant,
                colour=colour,
                size=size,
                on_hand=_to_int(request.POST.get("on_hand"), default=0),
                reserved=_to_int(request.POST.get("reserved"), default=0),
                is_active=True,
                buffer_min=_to_int(request.POST.get("buffer_min"), default=0),
                buffer_target=_to_int(request.POST.get("buffer_target"), default=0),
                buffer_max=_to_int(request.POST.get("buffer_max"), default=0),
            )
        else:
            existing.is_active = True
            existing.on_hand = _to_int(request.POST.get("on_hand"), default=0)
            existing.reserved = _to_int(request.POST.get("reserved"), default=0)
            existing.buffer_min = _to_int(request.POST.get("buffer_min"), default=0)
            existing.buffer_target = _to_int(request.POST.get("buffer_target"), default=0)
            existing.buffer_max = _to_int(request.POST.get("buffer_max"), default=0)
            existing.save(update_fields=["is_active", "on_hand", "reserved", "buffer_min", "buffer_target", "buffer_max", "updated_at"])
    except (ValueError, IntegrityError) as exc:
        messages.error(request, f"Could not create printed SKU: {exc}")
        return redirect("sku-manager")

    messages.success(request, "Printed SKU created.")
    return redirect("sku-manager")


@transaction.atomic
def _update_printed_group(request: HttpRequest) -> HttpResponse:
    """Update a consolidated printed-stock row across size columns."""
    design_name = request.POST.get("design_name", "").strip()
    colour = request.POST.get("colour", "").strip()
    variant = request.POST.get("variant", "").strip() or None
    if not design_name:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": "Design name is required."}, status=400)
        messages.error(request, "Design name is required.")
        return redirect("sku-manager")
    if not colour:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": "Colour is required."}, status=400)
        messages.error(request, "Colour is required.")
        return redirect("sku-manager")

    design, _ = Design.objects.get_or_create(name=design_name)
    try:
        for size in SIZE_COLUMNS:
            on_hand = _to_int(request.POST.get(f"on_hand_{size}"), default=0)

            sku = PrintedSKU.objects.select_for_update().filter(
                design=design,
                variant=variant,
                colour=colour,
                size=size,
            ).first()
            if sku is None:
                if on_hand == 0:
                    continue
                PrintedSKU.objects.create(
                    design=design,
                    variant=variant,
                    colour=colour,
                    size=size,
                    on_hand=on_hand,
                    is_active=True,
                )
            else:
                sku.is_active = True
                sku.on_hand = on_hand
                sku.save(update_fields=["is_active", "on_hand", "updated_at"])
    except (ValueError, IntegrityError) as exc:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": f"Could not update printed stock row: {exc}"}, status=400)
        messages.error(request, f"Could not update printed stock row: {exc}")
        return redirect("sku-manager")

    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": "Saved"}, status=200)

    messages.success(request, "Printed stock row updated.")
    return redirect("sku-manager")


@transaction.atomic
def _delete_printed_group(request: HttpRequest) -> HttpResponse:
    """Delete all printed SKU sizes for a design/variant/colour group and record payloads."""
    design_name = request.POST.get("design_name", "").strip()
    colour = request.POST.get("colour", "").strip()
    variant = request.POST.get("variant", "").strip() or None
    if not design_name or not colour:
        messages.error(request, "Design and colour are required for delete.")
        return redirect("sku-manager")

    design = Design.objects.filter(name=design_name).first()
    if design is None:
        messages.error(request, "Design not found.")
        return redirect("sku-manager")

    skus = list(
        PrintedSKU.objects.select_for_update().select_related("design").filter(
            design=design,
            variant=variant,
            colour=colour,
        )
    )
    if not skus:
        messages.error(request, "No printed stock rows found for this group.")
        return redirect("sku-manager")

    deleted = 0
    archived = 0
    for sku in skus:
        payload = {
            "design_name": sku.design.name,
            "variant": sku.variant,
            "colour": sku.colour,
            "size": sku.size,
            "on_hand": sku.on_hand,
            "reserved": sku.reserved,
            "buffer_min": sku.buffer_min,
            "buffer_target": sku.buffer_target,
            "buffer_max": sku.buffer_max,
        }
        DeletedInventoryItem.objects.create(
            record_type=DeletedInventoryItem.RecordType.PRINTED_SKU,
            source_model_id=str(sku.id),
            label=str(sku),
            payload=payload,
            deleted_by=request.user if request.user.is_authenticated else None,
        )
        try:
            sku.delete()
            deleted += 1
        except Exception as exc:
            # PROTECT constraints from historical orders prevent hard delete.
            # Archive instead to hide it from active stock management.
            sku.is_active = False
            sku.on_hand = 0
            if sku.reserved > 0:
                sku.reserved = 0
            sku.save(update_fields=["is_active", "on_hand", "reserved", "updated_at"])
            archived += 1

    if archived and deleted:
        messages.warning(request, f"Deleted {deleted} SKU size row(s). Archived {archived} protected row(s) linked to historical orders.")
    elif archived:
        messages.warning(request, f"Archived {archived} protected SKU size row(s) because they are linked to historical orders.")
    else:
        messages.success(request, f"Deleted {deleted} printed size row(s).")
    return redirect("sku-manager")


@transaction.atomic
def _link_blank_sku(request: HttpRequest) -> HttpResponse:
    """Link (or unlink) a blank SKU to a design+colour (via DesignAsset).
    
    This links at the design+colour level, so all sizes automatically share the same blank SKU.
    """
    design_name = request.POST.get("design_name", "").strip()
    colour = request.POST.get("colour", "").strip()
    plain_option = request.POST.get("plain_option", "").strip()

    # Find the design
    design = Design.objects.filter(name=design_name).first()
    if design is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "detail": "Design not found."}, status=404)
        messages.error(request, "Design not found.")
        return redirect("sku-manager")

    # Find or create the DesignAsset for this design+colour
    design_asset, created = DesignAsset.objects.get_or_create(
        design=design,
        colour=colour,
    )

    # Consolidated link stores fabric+colour intent at design-asset level.
    # We intentionally clear size-specific blank_sku to avoid incorrect per-size coupling.
    if plain_option:
        try:
            selected_fabric, selected_colour = [part.strip() for part in plain_option.split("|||", 1)]
        except ValueError:
            if _is_ajax(request):
                return JsonResponse({"ok": False, "detail": "Invalid plain SKU selection."}, status=400)
            messages.error(request, "Invalid plain SKU selection.")
            return redirect("sku-manager")

        if not BlankSKU.objects.filter(fabric=selected_fabric, colour=selected_colour).exists():
            if _is_ajax(request):
                return JsonResponse({"ok": False, "detail": "Selected plain SKU group not found."}, status=404)
            messages.error(request, "Selected plain SKU group not found.")
            return redirect("sku-manager")

        design_asset.blank_fabric = selected_fabric
        design_asset.blank_sku = None
        design_asset.save(update_fields=["blank_fabric", "blank_sku", "updated_at"])
    else:
        design_asset.blank_sku = None
        design_asset.save(update_fields=["blank_sku", "updated_at"])

    # Update all PrintedSKU of this design+colour to reference this design_asset
    # (For backward compatibility, also sync blank_sku on the PrintedSKU records)
    PrintedSKU.objects.filter(
        design=design,
        colour=colour,
    ).update(design_asset=design_asset, blank_sku=None)

    if _is_ajax(request):
        label = plain_option.replace("|||", " / ") if plain_option else "Unlinked"
        return JsonResponse({"ok": True, "message": f"Linked {design_name}/{colour}: {label}"})
    messages.success(request, f"Plain SKU link saved for {design_name}/{colour}.")
    return redirect("sku-manager")


@transaction.atomic
def _import_printed_skus(request: HttpRequest) -> HttpResponse:
    """Import printed SKUs from uploaded Excel sheet."""
    file = request.FILES.get("excel_file")
    if file is None:
        messages.error(request, "Choose an Excel file to import.")
        return redirect("sku-manager")

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        messages.error(request, f"Unable to read Excel file: {exc}")
        return redirect("sku-manager")

    headers: dict[str, int] = {}
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        messages.error(request, "Excel file is empty.")
        return redirect("sku-manager")

    for index, value in enumerate(first_row):
        header = str(value or "").strip().lower()
        if header:
            headers[header] = index

    required_headers = ["design_name", "colour", "on_hand", "reserved", "buffer_min", "buffer_target", "buffer_max"]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}")
        return redirect("sku-manager")

    created = 0
    updated = 0
    failed = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        design_name = str(row[headers["design_name"]] or "").strip()
        colour = str(row[headers["colour"]] or "").strip()
        if not design_name or not colour:
            continue

        try:
            design, _ = Design.objects.get_or_create(name=design_name)
            variant = str(row[headers.get("variant", -1)] or "").strip() or None if "variant" in headers else None
            size = str(row[headers.get("size", -1)] or "").strip() or None if "size" in headers else None

            queryset = PrintedSKU.objects.select_for_update().filter(
                design=design,
                variant=variant,
                colour=colour,
                size=size,
            )
            sku = queryset.first()
            if sku is None:
                PrintedSKU.objects.create(
                    design=design,
                    variant=variant,
                    colour=colour,
                    size=size,
                    on_hand=_to_int(row[headers["on_hand"]], default=0),
                    reserved=_to_int(row[headers["reserved"]], default=0),
                    is_active=True,
                    buffer_min=_to_int(row[headers["buffer_min"]], default=0),
                    buffer_target=_to_int(row[headers["buffer_target"]], default=0),
                    buffer_max=_to_int(row[headers["buffer_max"]], default=0),
                )
                created += 1
            else:
                sku.is_active = True
                sku.on_hand = _to_int(row[headers["on_hand"]], default=0)
                sku.reserved = _to_int(row[headers["reserved"]], default=0)
                sku.buffer_min = _to_int(row[headers["buffer_min"]], default=0)
                sku.buffer_target = _to_int(row[headers["buffer_target"]], default=0)
                sku.buffer_max = _to_int(row[headers["buffer_max"]], default=0)
                sku.save(update_fields=["is_active", "on_hand", "reserved", "buffer_min", "buffer_target", "buffer_max", "updated_at"])
                updated += 1
        except Exception as exc:
            failed += 1
            messages.warning(request, f"Row {row_number} failed: {exc}")

    messages.success(request, f"Import complete. Created: {created}, Updated: {updated}, Failed: {failed}")
    return redirect("sku-manager")
