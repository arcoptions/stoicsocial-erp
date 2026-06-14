"""
Management command to seed the database from the real Master Order Tracker Excel files.
Clears all existing sample data and loads real designs, inventory, and orders.
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime

try:
    import openpyxl
except ImportError:
    raise CommandError("openpyxl is required. Run: pip install openpyxl")

from core.models import (
    BlankSKU, Design, DesignAsset, Order, OrderLine,
    PrintBatch, PrintJob, PrintJobLine, PrintedSKU, Vendor,
)

EXCEL_PATH = Path("C:/Users/abhiramnarla/Downloads/Master Order Tracker (1).xlsx")

# Status mapping from Excel → Django model choices
EXCEL_TO_ORDER_STATUS = {
    "new":            Order.STATUS_NEW,
    "to be printed":  Order.STATUS_NEEDS_PRINTING,
    "existing":       Order.STATUS_NEEDS_PRINTING,
    "in printing":    Order.STATUS_IN_PRINTING,
    "ready ship":     Order.STATUS_READY_TO_SHIP,
    "shipped":        Order.STATUS_SHIPPED,
    "cancelled":      Order.STATUS_CANCELLED,
}

EXCEL_TO_LINE_STATUS = {
    "new":            OrderLine.STATUS_NEW,
    "to be printed":  OrderLine.STATUS_TO_BE_PRINTED,
    "existing":       OrderLine.STATUS_TO_BE_PRINTED,
    "in printing":    OrderLine.STATUS_IN_PRINTING,
    "ready ship":     OrderLine.STATUS_READY_SHIP,
    "shipped":        OrderLine.STATUS_SHIPPED,
    "cancelled":      OrderLine.STATUS_CANCELLED,
}

FULFILLMENT_TO_ORDER_STATUS = {
    "fulfilled": Order.STATUS_SHIPPED,
    "partial": Order.STATUS_READY_TO_SHIP,
    "partially_fulfilled": Order.STATUS_READY_TO_SHIP,
}

DELIVERY_TO_ORDER_STATUS = {
    "delivered": Order.STATUS_SHIPPED,
    "in_transit": Order.STATUS_SHIPPED,
    "out_for_delivery": Order.STATUS_SHIPPED,
}

FULFILLMENT_TO_LINE_STATUS = {
    "fulfilled": OrderLine.STATUS_SHIPPED,
    "partial": OrderLine.STATUS_READY_SHIP,
    "partially_fulfilled": OrderLine.STATUS_READY_SHIP,
}

DELIVERY_TO_LINE_STATUS = {
    "delivered": OrderLine.STATUS_SHIPPED,
    "in_transit": OrderLine.STATUS_SHIPPED,
    "out_for_delivery": OrderLine.STATUS_SHIPPED,
}

# Plain stock fabric map from tracker names
PLAIN_FABRIC_MAP = {
    "plain 180 gsm black": ("180 GSM", "Black"),
    "plain 180 gsm red":   ("180 GSM", "Red"),
    "plain 180 gsm white": ("180 GSM", "White"),
    "plain 180 gsm blue":  ("180 GSM", "Blue"),
    "plain 180 gsm navy":  ("180 GSM", "Navy"),
    "plain 180 gsm grey":  ("180 GSM", "Grey"),
}
SIZES = ["S", "M", "L", "XL", "2XL", "3XL"]


def _normalize_imported_statuses(
    excel_status: str,
    fulfillment_status: str,
    delivery_status: str,
) -> tuple[str, str]:
    """Resolve imported order and line statuses with fulfillment precedence.

    The tracker's workflow column can lag behind Shopify fulfillment state.
    When fulfillment or delivery shows a later stage, prefer that state.
    """
    normalized_excel_status = excel_status.strip().lower()
    normalized_fulfillment = fulfillment_status.strip().lower()
    normalized_delivery = delivery_status.strip().lower()

    order_status = EXCEL_TO_ORDER_STATUS.get(normalized_excel_status, Order.STATUS_NEW)
    line_status = EXCEL_TO_LINE_STATUS.get(normalized_excel_status, OrderLine.STATUS_NEW)

    if normalized_excel_status == "cancelled":
        return order_status, line_status

    delivery_override = DELIVERY_TO_ORDER_STATUS.get(normalized_delivery)
    fulfillment_override = FULFILLMENT_TO_ORDER_STATUS.get(normalized_fulfillment)
    if delivery_override:
        order_status = delivery_override
    elif fulfillment_override:
        order_status = fulfillment_override

    delivery_line_override = DELIVERY_TO_LINE_STATUS.get(normalized_delivery)
    fulfillment_line_override = FULFILLMENT_TO_LINE_STATUS.get(normalized_fulfillment)
    if delivery_line_override:
        line_status = delivery_line_override
    elif fulfillment_line_override:
        line_status = fulfillment_line_override

    return order_status, line_status


class Command(BaseCommand):
    help = "Load real data from Master Order Tracker Excel into BoldERP database."

    @transaction.atomic
    def handle(self, *args: Any, **options: Any) -> None:
        if not EXCEL_PATH.exists():
            raise CommandError(f"Excel file not found: {EXCEL_PATH}")

        self.stdout.write("Reading Excel file...")
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

        self.stdout.write("Clearing existing data...")
        self._clear_data()

        self.stdout.write("Loading plain stock (BlankSKUs)...")
        n_blanks = self._load_plain_stock(wb)
        self.stdout.write(f"  → {n_blanks} BlankSKU records")

        self.stdout.write("Loading designs and design assets...")
        n_designs = self._load_designs(wb)
        self.stdout.write(f"  → {n_designs} Design records")

        self.stdout.write("Loading printed stock (PrintedSKUs)...")
        n_printed = self._load_printed_stock(wb)
        self.stdout.write(f"  → {n_printed} PrintedSKU records")

        self.stdout.write("Loading vendor...")
        vendor = Vendor.objects.get_or_create(
            name="Knitwear",
            defaults={"contact": "vendor@knitwear.in", "is_active": True},
        )[0]

        self.stdout.write("Loading orders and order lines...")
        n_orders, n_lines = self._load_orders(wb)
        self.stdout.write(f"  → {n_orders} Orders, {n_lines} OrderLines")

        self.stdout.write("Creating receive queue from To Print sheet...")
        n_jobs = self._create_print_jobs(wb, vendor)
        self.stdout.write(f"  → {n_jobs} PrintJob(s) created")

        self.stdout.write(self.style.SUCCESS(
            f"\nDone! {n_blanks} blanks | {n_designs} designs | {n_printed} printed SKUs | "
            f"{n_orders} orders | {n_lines} lines | {n_jobs} print jobs"
        ))

    def _clear_data(self) -> None:
        from core.models import ReprintTask, StockMovement, WebhookEvent
        ReprintTask.objects.all().delete()
        StockMovement.objects.all().delete()
        WebhookEvent.objects.all().delete()
        PrintJobLine.objects.all().delete()
        PrintJob.objects.all().delete()
        PrintBatch.objects.all().delete()
        OrderLine.objects.all().delete()
        Order.objects.all().delete()
        PrintedSKU.objects.all().delete()
        DesignAsset.objects.all().delete()
        Design.objects.all().delete()
        BlankSKU.objects.all().delete()
        Vendor.objects.all().delete()

    def _load_plain_stock(self, wb: openpyxl.Workbook) -> int:
        ws = wb["Plain Stock"]
        rows = list(ws.iter_rows(values_only=True))
        count = 0
        for row in rows[1:]:
            product_name = str(row[0] or "").strip()
            if not product_name:
                continue
            fabric, colour = PLAIN_FABRIC_MAP.get(product_name.lower(), ("180 GSM", product_name))
            size_values = {SIZES[i]: int(row[i + 1] or 0) for i in range(len(SIZES)) if i + 1 < len(row)}
            for size, on_hand in size_values.items():
                if on_hand is None:
                    on_hand = 0
                BlankSKU.objects.update_or_create(
                    fabric=fabric,
                    colour=colour,
                    size=size,
                    defaults={
                        "on_hand": on_hand,
                        "reserved": 0,
                        "reorder_min": 10,
                        "reorder_target": 30,
                    },
                )
                count += 1
        return count

    def _load_designs(self, wb: openpyxl.Workbook) -> int:
        ws = wb["Designs Repository"]
        rows = list(ws.iter_rows(values_only=True))
        count = 0
        for row in rows[1:]:
            if not row[0]:
                continue
            name = str(row[0]).strip()
            product_type = str(row[1] or "Tshirt").strip()
            sub_category = str(row[2] or "Regular").strip()
            colour = str(row[3] or "Black").strip()
            design_link = str(row[4] or "").strip()
            mockup_link = str(row[5] or "").strip()

            design, _ = Design.objects.update_or_create(
                name=name,
                defaults={
                    "product_type": product_type,
                    "sub_category": sub_category,
                    "has_variants": False,
                    "variants": [],
                    "notes": "",
                },
            )
            DesignAsset.objects.update_or_create(
                design=design,
                colour=colour,
                defaults={
                    "artwork_url": design_link or "https://example.com/artwork.png",
                    "mockup_url": mockup_link or "https://example.com/mockup.png",
                    "blank_fabric": "180 GSM",
                },
            )
            count += 1
        return count

    def _load_printed_stock(self, wb: openpyxl.Workbook) -> int:
        ws = wb["Printed Stock V2"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip() for h in rows[0]]

        # Map size headers: S, M, L, XL, 2XL, 3XL
        size_col_map = {}
        for i, h in enumerate(headers):
            if h.strip() in SIZES:
                size_col_map[h.strip()] = i

        count = 0
        for row in rows[1:]:
            product = str(row[0] or "").strip()
            variant = str(row[1] or "").strip() if row[1] else None
            plain_stock_col = str(row[2] or "").strip()
            colour = "Black"
            if "white" in plain_stock_col.lower():
                colour = "White"
            elif "red" in plain_stock_col.lower():
                colour = "Red"
            elif "navy" in plain_stock_col.lower():
                colour = "Navy"
            elif "blue" in plain_stock_col.lower():
                colour = "Blue"

            if not product:
                continue

            # Get or create design
            design, _ = Design.objects.get_or_create(
                name=product,
                defaults={
                    "product_type": "Tshirt",
                    "sub_category": "Regular",
                    "has_variants": bool(variant),
                    "variants": [],
                },
            )

            for size, col_idx in size_col_map.items():
                qty = int(row[col_idx] or 0) if col_idx < len(row) else 0
                if qty <= 0:
                    continue
                PrintedSKU.objects.update_or_create(
                    design=design,
                    variant=variant or None,
                    colour=colour,
                    size=size,
                    defaults={
                        "on_hand": qty,
                        "reserved": 0,
                        "buffer_min": 3,
                        "buffer_target": 10,
                        "buffer_max": 30,
                    },
                )
                count += 1

        # Create at-least-one printed SKU for every design that has none yet
        for design in Design.objects.all():
            if not design.printed_skus.exists():
                for size in ["M", "L", "XL"]:
                    PrintedSKU.objects.get_or_create(
                        design=design,
                        variant=None,
                        colour="Black",
                        size=size,
                        defaults={
                            "on_hand": 0,
                            "reserved": 0,
                            "buffer_min": 5,
                            "buffer_target": 15,
                            "buffer_max": 40,
                        },
                    )
                    count += 1

        return count

    def _load_orders(self, wb: openpyxl.Workbook) -> tuple[int, int]:
        ws = wb["Master Order Tracker"]
        rows = list(ws.iter_rows(values_only=True))

        orders_by_no: dict[str, Order] = {}
        n_orders = 0
        n_lines = 0

        for row in rows[1:]:
            if not row[0]:
                continue
            shopify_id = str(int(row[0]))
            order_no = str(row[1] or "").strip()
            line_id = str(int(row[2])) if row[2] else f"line-{shopify_id}-{n_lines}"
            created_raw = str(row[3] or "").strip()
            customer_name = str(row[4] or "").strip()
            email = str(row[5] or "").strip()
            product = str(row[6] or "").strip()
            variant = str(row[7] or "").strip() or None
            size = str(row[8] or "").strip() or None
            qty = int(row[10] or 1)
            excel_status = str(row[11] or "new").strip().lower()
            fulfillment = str(row[12] or "").strip()
            tags_raw = str(row[13] or "").strip()
            delivery_status = str(row[14] or "").strip()

            tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []
            order_status, line_status = _normalize_imported_statuses(
                excel_status,
                fulfillment,
                delivery_status,
            )

            # Parse created_at
            created_at = timezone.now()
            if created_raw:
                try:
                    dt = parse_datetime(created_raw)
                    if dt:
                        created_at = dt if timezone.is_aware(dt) else timezone.make_aware(dt)
                except Exception:
                    pass

            # Create or get the parent order
            if order_no not in orders_by_no:
                order, created = Order.objects.update_or_create(
                    shopify_order_id=shopify_id,
                    defaults={
                        "order_no": order_no,
                        "customer_name": customer_name,
                        "email": email,
                        "tags": tags,
                        "status": order_status,
                        "shopify_fulfillment_status": fulfillment,
                        "shopify_delivery_status": delivery_status,
                        "raw_payload": {"source": "excel_import"},
                    },
                )
                orders_by_no[order_no] = order
                if created:
                    n_orders += 1
            else:
                order = orders_by_no[order_no]

            # Resolve printed SKU
            printed_sku = None
            if product:
                sku_qs = PrintedSKU.objects.filter(design__name__iexact=product)
                if size:
                    sized = sku_qs.filter(size__iexact=size).first()
                    printed_sku = sized or sku_qs.first()
                else:
                    printed_sku = sku_qs.first()

            # Create order line (idempotent by shopify_line_id)
            OrderLine.objects.update_or_create(
                shopify_line_id=line_id,
                defaults={
                    "order": order,
                    "product_name": product,
                    "variant": variant or "",
                    "size": size or "",
                    "quantity": qty,
                    "printed_sku": printed_sku,
                    "is_bundle": False,
                    "bundle_components": [],
                    "status": line_status,
                },
            )
            n_lines += 1

        # Recompute worst-case order statuses from their lines
        for order in Order.objects.prefetch_related("lines"):
            lines = list(order.lines.all())
            if lines:
                # Use the "least advanced" non-terminal status
                rank = {
                    Order.STATUS_NEW: 0,
                    Order.STATUS_NEEDS_PRINTING: 1,
                    Order.STATUS_IN_PRINTING: 2,
                    Order.STATUS_READY_TO_SHIP: 3,
                    Order.STATUS_SHIPPED: 4,
                    Order.STATUS_CANCELLED: 5,
                }
                line_order_statuses = []
                for ln in lines:
                    s = {
                        OrderLine.STATUS_NEW: Order.STATUS_NEW,
                        OrderLine.STATUS_TO_BE_PRINTED: Order.STATUS_NEEDS_PRINTING,
                        OrderLine.STATUS_IN_PRINTING: Order.STATUS_IN_PRINTING,
                        OrderLine.STATUS_READY_SHIP: Order.STATUS_READY_TO_SHIP,
                        OrderLine.STATUS_SHIPPED: Order.STATUS_SHIPPED,
                        OrderLine.STATUS_CANCELLED: Order.STATUS_CANCELLED,
                    }.get(ln.status, Order.STATUS_NEW)
                    line_order_statuses.append(s)
                non_terminal = [s for s in line_order_statuses if s not in {Order.STATUS_SHIPPED, Order.STATUS_CANCELLED}]
                best = min(non_terminal or line_order_statuses, key=lambda x: rank.get(x, 99))
                order.status = best
                order.save(update_fields=["status"])

        return n_orders, n_lines

    def _create_print_jobs(self, wb: openpyxl.Workbook, vendor: Vendor) -> int:
        ws = wb["To Print"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip() for h in rows[0]]

        # Find size column indices
        size_cols = {h: i for i, h in enumerate(headers) if h in SIZES}
        product_col = 0
        variant_col = 1 if "variant" in headers[1].lower() else None

        # Collect print job lines
        job_lines = []
        for row in rows[1:]:
            product = str(row[product_col] or "").strip()
            variant = str(row[variant_col] or "").strip() if variant_col else None
            if not product:
                continue
            for size, col in size_cols.items():
                qty = int(row[col] or 0) if col < len(row) else 0
                if qty <= 0:
                    continue
                sku = PrintedSKU.objects.filter(design__name__iexact=product, size__iexact=size).first()
                if not sku:
                    sku = PrintedSKU.objects.filter(design__name__iexact=product).first()
                blank = BlankSKU.objects.filter(colour="Black", size=size).first()
                job_lines.append((sku, blank, qty))

        if not job_lines:
            return 0

        batch = PrintBatch.objects.create(
            status=PrintBatch.STATUS_CONFIRMED,
            notes="Imported from To Print sheet",
        )
        job = PrintJob.objects.create(
            batch=batch,
            vendor=vendor,
            status=PrintJob.STATUS_SENT,
            sent_at=timezone.now(),
            notes="Imported from To Print sheet",
        )
        seen_skus: set = set()
        for sku, blank, qty in job_lines:
            if sku and str(sku.id) not in seen_skus:
                seen_skus.add(str(sku.id))
                PrintJobLine.objects.create(
                    print_job=job,
                    printed_sku=sku,
                    blank_sku=blank,
                    qty_sent=qty,
                )

        return 1
