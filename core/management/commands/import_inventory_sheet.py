from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import F

from core.models import BlankSKU, Design, PrintedSKU


def _to_int(value: Any) -> int:
    """Convert spreadsheet numeric values into safe integers."""
    if value in (None, ""):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(float(str(value).strip()))


def _normalize_size(value: Any) -> str:
    """Normalize size labels such as 'M ' or '2XL'."""
    return str(value or "").strip().upper()


def _normalize_fabric(product_name: str) -> str:
    """Normalize plain inventory product names into fabric labels."""
    raw = product_name.strip()
    lower = raw.lower()
    if lower.startswith("plain "):
        return raw[6:].strip()
    return raw


class Command(BaseCommand):
    help = "Import current plain/printed inventory from Excel sheets."

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument(
            "--file",
            required=True,
            help="Absolute path to workbook (e.g. C:/Users/.../Inventory Sheet.xlsx)",
        )
        parser.add_argument(
            "--plain-sheets",
            nargs="+",
            default=["Ajna Stock", "Knitwear Stock"],
            help="Sheet names containing plain inventory rows.",
        )
        parser.add_argument(
            "--printed-sheet",
            default="Printed Stock V2",
            help="Sheet name containing printed inventory rows.",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Replace existing on_hand values for imported keys (recommended for current snapshot imports).",
        )
        parser.add_argument(
            "--reset-existing",
            action="store_true",
            help="Delete all existing BlankSKU and PrintedSKU rows before import.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report only; do not write to database.",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        """Load plain and printed inventory snapshots from the provided workbook."""
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        workbook = openpyxl.load_workbook(str(path), data_only=True)

        plain_sheets: list[str] = options["plain_sheets"]
        printed_sheet: str = options["printed_sheet"]
        dry_run: bool = options["dry_run"]
        replace: bool = options["replace"]
        reset_existing: bool = options["reset_existing"]

        for sheet in plain_sheets + [printed_sheet]:
            if sheet not in workbook.sheetnames:
                raise CommandError(f"Sheet '{sheet}' not found in workbook. Available: {workbook.sheetnames}")

        plain_aggregate = self._parse_plain_inventory(workbook, plain_sheets)
        printed_rows = self._parse_printed_inventory(workbook, printed_sheet)

        self.stdout.write(self.style.HTTP_INFO("Inventory parse summary:"))
        self.stdout.write(f"  Plain keys: {len(plain_aggregate)}")
        self.stdout.write(f"  Printed rows with qty > 0: {len(printed_rows)}")

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry-run enabled: no DB writes performed."))
            return

        if reset_existing:
            self.stdout.write(self.style.WARNING("Resetting existing BlankSKU and PrintedSKU on_hand to reserved..."))
            with transaction.atomic():
                PrintedSKU.objects.update(on_hand=F("reserved"))
                BlankSKU.objects.update(on_hand=F("reserved"))

        blank_updates = self._apply_plain_inventory(plain_aggregate, replace=replace)
        printed_updates = self._apply_printed_inventory(printed_rows, replace=replace)

        self.stdout.write(self.style.SUCCESS("Inventory import complete."))
        self.stdout.write(f"  BlankSKU upserts: {blank_updates}")
        self.stdout.write(f"  PrintedSKU upserts: {printed_updates}")

    def _parse_plain_inventory(
        self,
        workbook: openpyxl.Workbook,
        sheet_names: list[str],
    ) -> dict[tuple[str, str, str], int]:
        """Aggregate plain stock across multiple supplier sheets by fabric/colour/size."""
        aggregate: dict[tuple[str, str, str], int] = defaultdict(int)

        for sheet_name in sheet_names:
            ws = workbook[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(value or "").strip().lower() for value in rows[0]]

            def idx(name: str) -> int:
                try:
                    return headers.index(name)
                except ValueError as exc:
                    raise CommandError(f"Column '{name}' missing in sheet '{sheet_name}'") from exc

            product_idx = idx("product name")
            size_idx = idx("size")
            color_idx = idx("color")
            qty_idx = idx("quantity")

            for row in rows[1:]:
                product = str(row[product_idx] or "").strip()
                size = _normalize_size(row[size_idx])
                color = str(row[color_idx] or "").strip().title()
                qty = _to_int(row[qty_idx])

                if not product or not size or not color:
                    continue
                if qty <= 0:
                    continue

                fabric = _normalize_fabric(product)
                key = (fabric, color, size)
                aggregate[key] += qty

        return dict(aggregate)

    def _parse_printed_inventory(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str,
    ) -> list[tuple[str, str, str, int]]:
        """Extract printed stock rows as design/colour/size/qty tuples."""
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value or "").strip() for value in rows[0]]
        normalized = [header.lower().strip() for header in headers]

        def idx(name: str) -> int:
            try:
                return normalized.index(name)
            except ValueError as exc:
                raise CommandError(f"Column '{name}' missing in sheet '{sheet_name}'") from exc

        color_idx = idx("color")
        product_idx = idx("product name")

        size_columns: dict[str, int] = {}
        for index, header in enumerate(headers):
            size = _normalize_size(header)
            if size in {"S", "M", "L", "XL", "2XL", "3XL"}:
                size_columns[size] = index

        if not size_columns:
            raise CommandError(f"No size columns found in sheet '{sheet_name}'")

        parsed: list[tuple[str, str, str, int]] = []
        for row in rows[1:]:
            design_name = str(row[product_idx] or "").strip()
            color = str(row[color_idx] or "").strip().title()
            if not design_name or not color:
                continue

            for size, col_index in size_columns.items():
                qty = _to_int(row[col_index]) if col_index < len(row) else 0
                if qty <= 0:
                    continue
                parsed.append((design_name, color, size, qty))

        return parsed

    def _apply_plain_inventory(self, plain_aggregate: dict[tuple[str, str, str], int], *, replace: bool) -> int:
        """Persist aggregated plain inventory snapshot to BlankSKU with row locks."""
        count = 0
        for (fabric, color, size), qty in plain_aggregate.items():
            with transaction.atomic():
                existing = (
                    BlankSKU.objects.select_for_update()
                    .filter(fabric=fabric, colour=color, size=size)
                    .first()
                )
                if existing is None:
                    BlankSKU.objects.create(
                        fabric=fabric,
                        colour=color,
                        size=size,
                        on_hand=qty,
                        reserved=0,
                        reorder_min=0,
                        reorder_target=0,
                    )
                else:
                    existing.on_hand = qty if replace else existing.on_hand + qty
                    existing.save(update_fields=["on_hand", "updated_at"])
            count += 1
            if count % 25 == 0:
                self.stdout.write(f"  Plain progress: {count}/{len(plain_aggregate)}")
        return count

    def _apply_printed_inventory(self, printed_rows: list[tuple[str, str, str, int]], *, replace: bool) -> int:
        """Persist printed inventory snapshot to PrintedSKU with row locks."""
        count = 0
        for design_name, color, size, qty in printed_rows:
            with transaction.atomic():
                design, _ = Design.objects.get_or_create(
                    name=design_name,
                    defaults={
                        "product_type": "Tshirt",
                        "sub_category": "Regular",
                        "material": "Cotton",
                        "fit": "Regular",
                        "has_variants": False,
                        "variants": [],
                        "notes": "Auto-created from printed inventory import",
                    },
                )

                existing = (
                    PrintedSKU.objects.select_for_update()
                    .filter(design=design, variant__isnull=True, colour=color, size=size)
                    .first()
                )
                if existing is None:
                    PrintedSKU.objects.create(
                        design=design,
                        variant=None,
                        colour=color,
                        size=size,
                        on_hand=qty,
                        reserved=0,
                        buffer_min=0,
                        buffer_target=0,
                        buffer_max=0,
                    )
                else:
                    existing.on_hand = qty if replace else existing.on_hand + qty
                    existing.save(update_fields=["on_hand", "updated_at"])
            count += 1
            if count % 25 == 0:
                self.stdout.write(f"  Printed progress: {count}/{len(printed_rows)}")

        return count
