"""
Script to inspect Excel file structure and prepare for import.
Run this to understand your current data format.
"""

import openpyxl
import json
from pathlib import Path


def inspect_excel(file_path: str) -> None:
    """Inspect Excel file structure and validate sheets."""
    
    file = Path(file_path)
    if not file.exists():
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return
    
    print("\n" + "="*80)
    print(f"📊 EXCEL FILE INSPECTION: {file.name}")
    print("="*80)
    
    # Check for required sheets
    required_sheets = ["vendors", "designs", "design_assets", "blank_skus", "printed_skus"]
    found_sheets = set(wb.sheetnames)
    
    print(f"\n✓ Found sheets: {', '.join(found_sheets)}")
    
    missing = set(required_sheets) - found_sheets
    if missing:
        print(f"⚠️  Missing sheets: {', '.join(missing)}")
    
    # Inspect each sheet
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            print(f"\n⚠️  Sheet '{sheet_name}' is empty")
            continue
        
        headers = rows[0]
        data_rows = len(rows) - 1
        
        print(f"\n📋 Sheet: {sheet_name}")
        print(f"   Columns: {headers}")
        print(f"   Rows: {data_rows}")
        
        # Show sample row
        if len(rows) > 1:
            print(f"   Sample row: {rows[1]}")
    
    print("\n" + "="*80)
    print("✅ Inspection complete. Check above for any ⚠️  warnings.\n")


def create_sample_workbook(output_path: str) -> None:
    """Create a sample Excel file with correct structure."""
    
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 1. Vendors
    ws = wb.create_sheet("vendors")
    ws.append(["name", "contact", "is_active"])
    ws.append(["PrintCo", "contact@printco.com", 1])
    ws.append(["FastPrint", "sales@fastprint.com", 1])
    ws.append(["QuickPrint", "info@quickprint.com", 1])
    
    # 2. Designs
    ws = wb.create_sheet("designs")
    ws.append(["product_type", "sub_category", "variants", "notes"])
    ws.append(["T-Shirt", "Basic", '["XS", "S", "M", "L", "XL"]', "Classic crew neck"])
    ws.append(["T-Shirt", "Premium", '["S", "M", "L", "XL", "XXL"]', "Premium 100% cotton"])
    ws.append(["Hoodie", "Basic", '["S", "M", "L", "XL"]', "Standard hoodie"])
    
    # 3. Design Assets
    ws = wb.create_sheet("design_assets")
    ws.append(["design_product_type", "design_sub_category", "colour", "artwork_url", "mockup_url", "blank_fabric"])
    ws.append(["T-Shirt", "Basic", "White", "https://example.com/artwork1.png", "https://example.com/mockup1.png", "Cotton"])
    ws.append(["T-Shirt", "Basic", "Black", "https://example.com/artwork2.png", "https://example.com/mockup2.png", "Cotton"])
    ws.append(["T-Shirt", "Premium", "Red", "https://example.com/artwork3.png", "https://example.com/mockup3.png", "Premium Cotton"])
    
    # 4. Blank SKUs
    ws = wb.create_sheet("blank_skus")
    ws.append(["fabric", "colour", "size", "on_hand", "reserved", "reorder_min", "reorder_target"])
    ws.append(["Cotton", "White", "S", 100, 0, 20, 50])
    ws.append(["Cotton", "White", "M", 150, 0, 25, 60])
    ws.append(["Cotton", "White", "L", 120, 0, 20, 50])
    ws.append(["Cotton", "Black", "S", 80, 0, 20, 50])
    ws.append(["Cotton", "Black", "M", 110, 0, 25, 60])
    ws.append(["Cotton", "Black", "L", 90, 0, 20, 50])
    
    # 5. Printed SKUs
    ws = wb.create_sheet("printed_skus")
    ws.append(["design_product_type", "design_sub_category", "variant", "colour", "size", "on_hand", "reserved", "buffer_min", "buffer_target", "buffer_max"])
    ws.append(["T-Shirt", "Basic", "XS", "White", "XS", 30, 0, 5, 15, 50])
    ws.append(["T-Shirt", "Basic", "S", "White", "S", 45, 0, 10, 25, 75])
    ws.append(["T-Shirt", "Basic", "M", "White", "M", 60, 0, 10, 30, 100])
    ws.append(["T-Shirt", "Basic", "XS", "Black", "XS", 25, 0, 5, 15, 50])
    ws.append(["T-Shirt", "Basic", "S", "Black", "S", 40, 0, 10, 25, 75])
    
    wb.save(output_path)
    print(f"✅ Sample workbook created: {output_path}")


if __name__ == "__main__":
    import sys
    
    print("""
    BoldERP Excel File Validator
    =============================
    
    Usage:
      python validate_excel.py <path_to_excel_file>
      
    Example:
      python validate_excel.py "c:\\path\\to\\Master Order Tracker.xlsx"
    """)
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        inspect_excel(file_path)
    else:
        print("Creating sample workbook...")
        create_sample_workbook("sample_master_data.xlsx")
        print("\nNow inspect it:")
        print('  python validate_excel.py "sample_master_data.xlsx"')
